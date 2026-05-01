"""Built-in DocFrame adapters."""

from __future__ import annotations

import csv
import logging
import uuid
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader

from .models import DocumentChunk, DocumentResult, ProcessingOptions
from .registry import AdapterRegistry, DocumentAdapter
from .utils import build_metadata, split_text

for logger_name in ("pypdf", "pypdf._cmap"):
    logger = logging.getLogger(logger_name)
    logger.setLevel(logging.CRITICAL)
    logger.propagate = False
    logger.addHandler(logging.NullHandler())


def chunk_id() -> str:
    """Return a chunk UUID."""

    return str(uuid.uuid4())


WORD_NAMESPACE = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


class CsvAdapter(DocumentAdapter):
    """CSV adapter using Python's standard library parser."""

    extensions = frozenset({".csv"})

    def process(self, path: Path, options: ProcessingOptions) -> DocumentResult:
        metadata = build_metadata(path)
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(4096)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|") if sample else csv.excel
            except csv.Error:
                dialect = csv.excel
            try:
                has_header = csv.Sniffer().has_header(sample) if sample else True
            except csv.Error:
                has_header = True
            reader = csv.reader(handle, dialect)
            rows = [row for row in reader if any(cell.strip() for cell in row)]

        if not rows:
            metadata.row_count = 0
            return DocumentResult(document_id=metadata.sha256, metadata=metadata, chunks=[])

        headers = (
            [str(value) for value in rows[0]]
            if has_header
            else [f"column_{index + 1}" for index in range(len(rows[0]))]
        )
        data_rows = rows[1:] if rows and has_header else rows
        normalized = [
            {
                headers[index] if index < len(headers) else f"column_{index + 1}": value
                for index, value in enumerate(row)
            }
            for row in data_rows[: options.max_table_rows]
        ]
        metadata.row_count = len(data_rows)
        chunks = [
            DocumentChunk(
                id=chunk_id(),
                type="table",
                rows=normalized,
                source=metadata.filename,
                metadata={"truncated": len(data_rows) > options.max_table_rows},
            )
        ]
        return DocumentResult(document_id=metadata.sha256, metadata=metadata, chunks=chunks)


class ExcelAdapter(DocumentAdapter):
    """Excel adapter for `.xlsx` and `.xlsm` workbooks."""

    extensions = frozenset({".xlsx", ".xlsm"})

    def process(self, path: Path, options: ProcessingOptions) -> DocumentResult:
        metadata = build_metadata(path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        chunks: list[DocumentChunk] = []
        total_rows = 0

        try:
            metadata.sheet_count = len(workbook.worksheets)
            for sheet in workbook.worksheets:
                rows_iter = sheet.iter_rows(values_only=True)
                first = next(rows_iter, None)
                if first is None:
                    continue
                headers = [
                    str(value) if value is not None else f"column_{index + 1}"
                    for index, value in enumerate(first)
                ]
                rows: list[dict[str, Any]] = []
                sheet_rows = 0
                for raw_row in rows_iter:
                    sheet_rows += 1
                    if len(rows) >= options.max_table_rows:
                        continue
                    rows.append(
                        {
                            headers[index]: value
                            for index, value in enumerate(raw_row)
                            if index < len(headers)
                        }
                    )
                total_rows += sheet_rows
                chunks.append(
                    DocumentChunk(
                        id=chunk_id(),
                        type="table",
                        rows=rows,
                        sheet=sheet.title,
                        source=metadata.filename,
                        metadata={"truncated": sheet_rows > options.max_table_rows},
                    )
                )
        finally:
            workbook.close()

        metadata.row_count = total_rows
        return DocumentResult(document_id=metadata.sha256, metadata=metadata, chunks=chunks)


def _word_text(element: ElementTree.Element) -> str:
    """Extract visible text from a WordprocessingML element."""

    parts: list[str] = []
    for node in element.iter():
        if node.tag == f"{WORD_NAMESPACE}t" and node.text:
            parts.append(node.text)
        elif node.tag == f"{WORD_NAMESPACE}tab":
            parts.append("\t")
        elif node.tag == f"{WORD_NAMESPACE}br":
            parts.append("\n")
    return "".join(parts)


def _word_table_rows(table: ElementTree.Element) -> list[list[str]]:
    """Extract table rows from a WordprocessingML table element."""

    rows: list[list[str]] = []
    for row in table.findall(f"{WORD_NAMESPACE}tr"):
        cells: list[str] = []
        for cell in row.findall(f"{WORD_NAMESPACE}tc"):
            paragraphs = [
                _word_text(paragraph).strip()
                for paragraph in cell.findall(f".//{WORD_NAMESPACE}p")
            ]
            cells.append("\n".join(paragraph for paragraph in paragraphs if paragraph))
        if any(value.strip() for value in cells):
            rows.append(cells)
    return rows


def _extract_word_chunks(path: Path, metadata_filename: str, options: ProcessingOptions) -> list[DocumentChunk]:
    """Extract paragraph and table chunks from an OOXML Word package."""

    with zipfile.ZipFile(path) as archive:
        document_xml = archive.read("word/document.xml")

    root = ElementTree.fromstring(document_xml)
    body = root.find(f"{WORD_NAMESPACE}body")
    if body is None:
        return []

    chunks: list[DocumentChunk] = []
    paragraphs: list[str] = []
    table_index = 0

    for child in body:
        if child.tag == f"{WORD_NAMESPACE}p":
            text = _word_text(child).strip()
            if text:
                paragraphs.append(text)
        elif child.tag == f"{WORD_NAMESPACE}tbl":
            rows = _word_table_rows(child)
            if not rows:
                continue
            table_index += 1
            headers = rows[0]
            normalized = [
                {
                    (
                        headers[index]
                        if index < len(headers) and headers[index]
                        else f"column_{index + 1}"
                    ): value
                    for index, value in enumerate(row)
                }
                for row in rows[1 : options.max_table_rows + 1]
            ]
            chunks.append(
                DocumentChunk(
                    id=chunk_id(),
                    type="table",
                    rows=normalized,
                    source=metadata_filename,
                    metadata={
                        "table_index": table_index,
                        "truncated": max(0, len(rows) - 1) > options.max_table_rows,
                    },
                )
            )

    text = "\n".join(paragraphs)
    for index, part in enumerate(split_text(text, max_chars=options.max_chars_per_text_chunk)):
        chunks.append(
            DocumentChunk(
                id=chunk_id(),
                type="text",
                text=part,
                source=metadata_filename,
                metadata={"part": index + 1},
            )
        )

    return chunks


class DocxAdapter(DocumentAdapter):
    """DOCX adapter that extracts paragraphs and tables."""

    extensions = frozenset({".docx"})

    def process(self, path: Path, options: ProcessingOptions) -> DocumentResult:
        metadata = build_metadata(path)
        chunks = _extract_word_chunks(path, metadata.filename, options)
        return DocumentResult(document_id=metadata.sha256, metadata=metadata, chunks=chunks)


class LegacyDocAdapter(DocumentAdapter):
    """`.doc` adapter with OOXML recovery and safe legacy fallback.

    Many public corpora contain OOXML Word packages with a `.doc` extension.
    DocFrame extracts those directly. True historical binary Word files need
    platform-specific conversion tools or a dedicated parser, so they fall back
    to metadata-only output with a clear warning.
    """

    extensions = frozenset({".doc"})

    def process(self, path: Path, options: ProcessingOptions) -> DocumentResult:
        metadata = build_metadata(path)
        try:
            chunks = _extract_word_chunks(path, metadata.filename, options)
        except Exception as exc:
            metadata.extra["word_ooxml_error"] = f"{type(exc).__name__}: {exc}"
            chunks = [
                DocumentChunk(
                    id=chunk_id(),
                    type="metadata",
                    source=metadata.filename,
                    metadata={
                        "format": "legacy_word_binary",
                        "extraction": "metadata_only",
                    },
                )
            ]
            return DocumentResult(
                document_id=metadata.sha256,
                metadata=metadata,
                chunks=chunks,
                warnings=[
                    "Legacy .doc text extraction is not enabled. Convert to .docx or register a custom adapter."
                ],
            )

        warnings = [
            "File has a .doc extension but contains an OOXML Word package; processed with the DOCX extractor."
        ]
        return DocumentResult(
            document_id=metadata.sha256,
            metadata=metadata,
            chunks=chunks,
            warnings=warnings,
        )


class PdfAdapter(DocumentAdapter):
    """PDF adapter using pypdf text extraction."""

    extensions = frozenset({".pdf"})

    def process(self, path: Path, options: ProcessingOptions) -> DocumentResult:
        metadata = build_metadata(path)
        reader = PdfReader(str(path))
        metadata.page_count = len(reader.pages)
        chunks: list[DocumentChunk] = []
        warnings: list[str] = []

        if reader.is_encrypted:
            warnings.append("PDF is encrypted; text extraction may be unavailable.")

        for index, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            if not text.strip():
                continue
            for part_index, part in enumerate(
                split_text(text, max_chars=options.max_chars_per_text_chunk),
                start=1,
            ):
                chunks.append(
                    DocumentChunk(
                        id=chunk_id(),
                        type="text",
                        text=part,
                        page=index,
                        source=metadata.filename,
                        metadata={"part": part_index},
                    )
                )

        return DocumentResult(
            document_id=metadata.sha256,
            metadata=metadata,
            chunks=chunks,
            warnings=warnings,
        )


class ImageAdapter(DocumentAdapter):
    """Image adapter for PNG and JPEG metadata.

    OCR is intentionally an extension point. The built-in adapter normalizes
    image metadata and emits an image chunk so a vision/OCR provider can attach
    text later in the pipeline.
    """

    extensions = frozenset({".jpg", ".jpeg", ".png"})

    def process(self, path: Path, options: ProcessingOptions) -> DocumentResult:
        metadata = build_metadata(path)
        chunks: list[DocumentChunk] = []
        warnings = ["No OCR provider configured; emitted image metadata only."]

        with Image.open(path) as image:
            metadata.width = image.width
            metadata.height = image.height
            metadata.extra["mode"] = image.mode
            metadata.extra["format"] = image.format
            chunks.append(
                DocumentChunk(
                    id=chunk_id(),
                    type="image",
                    source=metadata.filename,
                    metadata={
                        "width": image.width,
                        "height": image.height,
                        "mode": image.mode,
                        "format": image.format,
                    },
                )
            )

        return DocumentResult(
            document_id=metadata.sha256,
            metadata=metadata,
            chunks=chunks,
            warnings=warnings,
        )


def adapter_registry() -> AdapterRegistry:
    """Return a registry with all built-in adapters installed."""

    registry = AdapterRegistry()
    registry.register(PdfAdapter())
    registry.register(LegacyDocAdapter())
    registry.register(DocxAdapter())
    registry.register(CsvAdapter())
    registry.register(ExcelAdapter())
    registry.register(ImageAdapter())
    return registry
